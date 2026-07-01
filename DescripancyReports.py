import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font
from db_credentials import get_sql_connection


class DownloadDiscrepancyReports:

    def __init__(self, root):

        self.root = root
        self.root.title("Download Descripancy Reports")
        self.root.geometry("1200x700")

        self.report_data = {}

        self.create_ui()
        self.load_reports()

    # ======================================================
    # UI
    # ======================================================

    def create_ui(self):

        # Header
        lbl_header = tk.Label(
            self.root,
            text="Download Descripancy Reports",
            font=("Arial", 18, "bold")
        )
        lbl_header.pack(pady=10)

        # Main Container
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill="both", expand=True)

        # ==================================================
        # LEFT PANEL (30%)
        # ==================================================
        left_frame = tk.Frame(
            main_frame,
            width=400
        )

        left_frame.pack(
            side="left",
            fill="y",
            padx=10,
            pady=10
        )

        left_frame.pack_propagate(False)

        lbl_list = tk.Label(
            left_frame,
            text="Download Descripancy Report",
            font=("Arial", 13, "bold")
        )

        lbl_list.pack(anchor="w", pady=(0, 5))

        # Listbox + Scrollbar Frame
        list_frame = tk.Frame(left_frame)
        list_frame.pack(fill="both", expand=True)

        # Scrollbar
        report_scrollbar = tk.Scrollbar(
            list_frame,
            orient="vertical"
        )

        report_scrollbar.pack(
            side="right",
            fill="y"
        )

        # Report List
        self.lst_reports = tk.Listbox(
            list_frame,
            font=("Arial", 13),
            yscrollcommand=report_scrollbar.set
        )

        self.lst_reports.pack(
            side="left",
            fill="both",
            expand=True
        )

        report_scrollbar.config(
            command=self.lst_reports.yview
        )

        self.lst_reports.bind(
            "<<ListboxSelect>>",
            self.report_selected
        )

        # Mouse Wheel Support
        self.lst_reports.bind(
            "<MouseWheel>",
            lambda e: self.lst_reports.yview_scroll(
                int(-1 * (e.delta / 120)),
                "units"
            )
        )

        # ==================================================
        # RIGHT PANEL
        # ==================================================
        right_frame = tk.Frame(main_frame)

        right_frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=10,
            pady=10
        )

        lbl_param = tk.Label(
            right_frame,
            text="Parameters",
            font=("Arial", 13, "bold")
        )

        lbl_param.pack(anchor="w")

        # Parameter Grid
        self.grid = ttk.Treeview(
            right_frame,
            columns=("Param Name", "Param Value"),
            show="headings",
            height=25
        )

        self.grid.heading(
            "Param Name",
            text="Param Name"
        )

        self.grid.heading(
            "Param Value",
            text="Param Value"
        )

        self.grid.column(
            "Param Name",
            width=250,
            anchor="w"
        )

        self.grid.column(
            "Param Value",
            width=300,
            anchor="w"
        )

        self.grid.pack(
            fill="both",
            expand=True
        )

        self.grid.bind(
            "<Double-1>",
            self.edit_param_value
        )

        # Download Button
        button_frame = tk.Frame(self.root)
        button_frame.pack(
            fill="x",
            pady=10
        )

        btn_download = tk.Button(
            button_frame,
            text="Download Excel",
            width=20,
            font=("Arial", 11, "bold"),
            bg="#1976D2",
            fg="white",
            command=self.download_excel
        )

        btn_download.pack()

    # ======================================================
    # DATABASE CONNECTION
    # ======================================================

    def get_connection(self):

        return get_sql_connection()

    # ======================================================
    # LOAD REPORTS
    # ======================================================

    def load_reports(self):

        try:

            connection = self.get_connection()

            cursor = connection.cursor()

            cursor.execute("""
                SELECT
                    ReportName,
                    ProcedureName,
                    Parametres
                FROM ExportReport
                ORDER BY ReportName
            """)

            self.lst_reports.delete(0, tk.END)
            self.report_data.clear()

            for row in cursor.fetchall():

                report_name = str(row.ReportName)

                self.report_data[report_name] = {
                    "ProcedureName": row.ProcedureName,
                    "Parametres": row.Parametres
                }

                self.lst_reports.insert(
                    tk.END,
                    report_name
                )

            connection.close()

        except Exception as ex:

            messagebox.showerror(
                "Error",
                str(ex)
            )

    # ======================================================
    # REPORT SELECTED
    # ======================================================

    def report_selected(self, event=None):

        selected = self.lst_reports.curselection()

        if not selected:
            return

        report_name = self.lst_reports.get(
            selected[0]
        )

        parameter_string = self.report_data[
            report_name
        ]["Parametres"]

        self.load_parameters(parameter_string)

    # ======================================================
    # LOAD PARAMETERS
    # ======================================================

    def load_parameters(self, parameter_string):

        self.grid.delete(*self.grid.get_children())

        if not parameter_string:
            return

        parameters = str(parameter_string).split(",")

        for parameter in parameters:

            self.grid.insert(
                "",
                "end",
                values=(
                    parameter.strip(),
                    ""
                )
            )

    # ======================================================
    # EDIT PARAMETER VALUE
    # ======================================================

    def edit_param_value(self, event):

        column = self.grid.identify_column(event.x)

        if column != "#2":
            return

        rowid = self.grid.identify_row(event.y)

        if not rowid:
            return

        x, y, width, height = self.grid.bbox(
            rowid,
            column
        )

        values = list(
            self.grid.item(
                rowid,
                "values"
            )
        )

        entry = tk.Entry(
            self.grid,
            font=("Arial", 10)
        )

        entry.place(
            x=x,
            y=y,
            width=width,
            height=height
        )

        entry.insert(
            0,
            values[1]
        )

        entry.focus()

        def save_value(event=None):

            values[1] = entry.get()

            self.grid.item(
                rowid,
                values=values
            )

            entry.destroy()

        entry.bind(
            "<Return>",
            save_value
        )

        entry.bind(
            "<FocusOut>",
            save_value
        )

    # ======================================================
    # EXECUTE PROCEDURE
    # ======================================================

    def execute_procedure(self, procedure_name):

        connection = self.get_connection()

        cursor = connection.cursor()

        parameter_values = []

        for item in self.grid.get_children():

            values = self.grid.item(
                item,
                "values"
            )

            parameter_values.append(
                values[1]
            )

        placeholders = ",".join(
            ["?"] * len(parameter_values)
        )

        if parameter_values:

            sql = f"EXEC {procedure_name} {placeholders}"

            cursor.execute(
                sql,
                parameter_values
            )

        else:

            cursor.execute(
                f"EXEC {procedure_name}"
            )

        columns = [
            column[0]
            for column in cursor.description
        ]

        data = cursor.fetchall()

        connection.close()

        return columns, data

    # ======================================================
    # EXPORT EXCEL
    # ======================================================

    def export_to_excel(
            self,
            columns,
            data,
            report_name):

        file_name = filedialog.asksaveasfilename(
            title="Save Excel Report",
            defaultextension=".xlsx",
            initialfile=f"{report_name}.xlsx",
            filetypes=[
                (
                    "Excel Files",
                    "*.xlsx"
                )
            ]
        )

        if not file_name:
            return

        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = "Report"

        # Headers
        for col_no, column_name in enumerate(
                columns,
                start=1):

            cell = worksheet.cell(
                row=1,
                column=col_no
            )

            cell.value = column_name
            cell.font = Font(bold=True)

        # Data
        row_no = 2

        for row in data:

            for col_no, value in enumerate(
                    row,
                    start=1):

                worksheet.cell(
                    row=row_no,
                    column=col_no
                ).value = value

            row_no += 1

        workbook.save(file_name)

        messagebox.showinfo(
            "Success",
            "Excel file downloaded successfully."
        )

    # ======================================================
    # DOWNLOAD EXCEL
    # ======================================================

    def download_excel(self):

        try:

            selected = self.lst_reports.curselection()

            if not selected:

                messagebox.showwarning(
                    "Warning",
                    "Please select a report."
                )

                return

            report_name = self.lst_reports.get(
                selected[0]
            )

            procedure_name = self.report_data[
                report_name
            ]["ProcedureName"]

            columns, data = self.execute_procedure(
                procedure_name
            )

            if len(data) == 0:

                messagebox.showinfo(
                    "Information",
                    "No data found."
                )

                return

            self.export_to_excel(
                columns,
                data,
                report_name
            )

        except Exception as ex:

            messagebox.showerror(
                "Error",
                str(ex)
            )


# ==========================================================
# MAIN
# ==========================================================
if __name__ == "__main__":

    root = tk.Tk()

    app = DownloadDiscrepancyReports(root)

    root.mainloop()
